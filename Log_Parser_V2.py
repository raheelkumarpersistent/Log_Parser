import os
from openpyxl import Workbook

path = "E:/Log Project/Files"  # Enter name of the folder where the log files are

# we shall store all the file names in this list
filelist = []

for root, dirs, files in os.walk(path):
    for file in files:
        # append the file name to the list
        filelist.append(os.path.join(root, file))

# print all the file names
for name in filelist:
    # print(name)
    wb = Workbook()
    sh1 = wb.active
    sh1.cell(1, 1, 'File name')
    sh1.cell(1, 2, 'Time Stamp')
    sh1.cell(1, 3, 'Transaction')
    rc = 2  # rownumber in excel
    cc = 1  # colnumber in excel
    fp1 = open(name, "r")  # Source file to read from
    data = fp1.readlines()
    # print(os.path.basename(fp1.name))  #to fetch the name of the source file
    flag = 0
    c = 0
    for line in data:
        if 'CB_To_SQL' in line:
            cc = 1
            c += 1
            each = list(map(str, line.split("] ")))
            for i in each:
                if each.index(i) == 0:
                    att = list(map(str, i.split(" ")))
                    # print(att[0:6], end="/// \n")
                    sh1.cell(rc, cc, os.path.basename(fp1.name))
                    cc += 1
                    sh1.cell(rc, cc, att[0] + ' ' + att[1] + ' ' + att[2] + ' ' + att[3] + ' ' + att[4]+']')
                    cc += 1
                    flag = 1
        elif ');' in line:
            flag = 0
            # print(line)
            sh1.cell(rc, cc, line)
            rc += 2
        if flag == 1:
            # print(line)
            sh1.cell(rc, cc, line)
            rc += 1
    fp1.close()
    # wb.save('New_Report.xlsx')
    wb.save(name[:-4] + '.xls')

    # wb.save(f'E:/Log Project/output/{name[:-4]}.xls')
